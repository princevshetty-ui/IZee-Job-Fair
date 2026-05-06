import io
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INDIGO = "FF6366F1"
INDIGO_DARK = "FF4F46E5"
WHITE = "FFFFFFFF"
DARK_BG = "FF0D0D1A"
SLATE = "FF334155"
GREEN = "FF10B981"
YELLOW = "FFF59E0B"
BLUE = "FF0EA5E9"
PURPLE = "FF8B5CF6"

ATTENDEE_COLS = [
    ("SID", "sid", 14),
    ("Full Name", "full_name", 24),
    ("Email", "email", 30),
    ("Phone", "phone", 16),
    ("City", "city", 16),
    ("State", "state", 18),
    ("College Name", "college_name", 28),
    ("Academic Level", "academic_level", 16),
    ("Stream", "stream", 22),
    ("Stream Other", "stream_other", 20),
    ("Attendee Type", "attendee_type", 16),
    ("MBA Specialization", "mba_specialization", 22),
    ("Graduation College", "graduation_college", 28),
    ("Graduation Stream", "graduation_stream", 20),
    ("Graduation Year", "graduation_year", 14),
    ("Company Name", "company_name", 24),
    ("Designation", "designation", 20),
    ("Experience Years", "experience_years", 16),
    ("Principal Name", "principal_name", 22),
    ("Principal Email", "principal_email", 28),
    ("Coordinator Name", "coordinator_name", 22),
    ("Coordinator Phone", "coordinator_phone", 18),
    ("Coordinator Email", "coordinator_email", 28),
    ("Reg Type", "reg_type", 12),
    ("Status", "status", 12),
    ("Attended", None, 10),
    ("Attended At", "attended_at", 22),
    ("Created At", "created_at", 22),
]

VOLUNTEER_COLS = [
    ("Name", "name", 24),
    ("Email", "email", 30),
    ("Phone", "phone", 16),
    ("Role", "role", 20),
    ("Created At", "created_at", 22),
]


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    s = Side(style="thin", color="FF1E293B")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_sheet(ws, columns: list, rows: List[Dict], tab_color: str = INDIGO):
    ws.sheet_properties.tabColor = tab_color[2:]

    header_font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    header_fill = _header_fill(INDIGO)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    even_fill = PatternFill("solid", fgColor="FF0F0F1F")
    odd_fill = PatternFill("solid", fgColor="FF0D0D1A")
    data_font = Font(color="FFCBD5E1", size=9, name="Calibri")
    data_align = Alignment(vertical="center")

    for row_idx, record in enumerate(rows, start=2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, (header, key, _) in enumerate(columns, start=1):
            if key is None and header == "Attended":
                value = "Yes" if record.get("attended") else "No"
            elif key is not None:
                value = record.get(key, "") or ""
            else:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = data_align
            cell.border = _thin_border()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(len(rows) + 1, 1)}"


def export_master_excel(
    all_attendees: List[Dict],
    volunteers: List[Dict],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    pre = [a for a in all_attendees if a.get("reg_type") == "pre"]
    onspot = [a for a in all_attendees if a.get("reg_type") == "onspot"]
    attended = [a for a in all_attendees if a.get("attended")]
    students = [a for a in all_attendees if (a.get("attendee_type") or "").lower() == "student"]
    freshers = [a for a in all_attendees if (a.get("attendee_type") or "").lower() == "fresher"]
    professionals = [a for a in all_attendees if (a.get("attendee_type") or "").lower() == "professional"]

    sheets = [
        ("All Registrations", INDIGO, all_attendees, ATTENDEE_COLS),
        ("Pre-Registered", BLUE, pre, ATTENDEE_COLS),
        ("On-Spot", GREEN, onspot, ATTENDEE_COLS),
        ("Attended", "FF2DD4BF", attended, ATTENDEE_COLS),
        ("Students", GREEN, students, ATTENDEE_COLS),
        ("Freshers", PURPLE, freshers, ATTENDEE_COLS),
        ("Professionals", YELLOW, professionals, ATTENDEE_COLS),
        ("Volunteers", "FFFBBF24", volunteers, VOLUNTEER_COLS),
    ]

    for name, color, data, cols in sheets:
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, cols, data, tab_color=color)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pre_excel(attendees: List[Dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    students = [a for a in attendees if (a.get("attendee_type") or "").lower() == "student"]
    freshers = [a for a in attendees if (a.get("attendee_type") or "").lower() == "fresher"]
    professionals = [a for a in attendees if (a.get("attendee_type") or "").lower() == "professional"]

    for name, color, data in [
        ("All Pre-Registered", INDIGO, attendees),
        ("Students", GREEN, students),
        ("Freshers", PURPLE, freshers),
        ("Professionals", YELLOW, professionals),
    ]:
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, ATTENDEE_COLS, data, tab_color=color)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_onspot_excel(attendees: List[Dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="On-Spot Registrations")
    _write_sheet(ws, ATTENDEE_COLS, attendees, tab_color=GREEN)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_volunteers_excel(volunteers: List[Dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Volunteers")
    _write_sheet(ws, VOLUNTEER_COLS, volunteers, tab_color=YELLOW)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_attended_excel(attendees: List[Dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Attended")
    _write_sheet(ws, ATTENDEE_COLS, attendees, tab_color="FF2DD4BF")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
